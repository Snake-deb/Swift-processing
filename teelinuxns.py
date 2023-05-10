import json
from os import listdir
import os.path as path
import re
import sqlite3
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED
from selenium.common.exceptions import TimeoutException

def list_files(directory, extension):
    return (f for f in listdir(directory) if f.endswith('.' + extension))

def check_vals(key):
    for x in range(len(dict_keys)):
        if key in dict_keys[x]:
            return dict_keys[x][key];
    return -1

def check_keys(key):
    for x in range(len(dict_keys)):
        if key in dict_keys[x]:
            return x
        else:
            for y in dict_keys[x].keys():
                if key == dict_keys[x][y]:
                    return x
    return -1

def check_rem(key):
    for x in range(len(dict_keys)):
        for y in dict_keys[x].keys():
                if key == dict_keys[x][y]:
                    return x
    return -1

def temp_driver(vgst):
    global driver_state
    global https
    global broken
    global drivercount
    drivercount += 1
    if drivercount > 2:
        drivercount = 0
        broken.append(https[0])
        https.pop(0)
        driver = webdriver.Firefox(executable_path="/mnt/d/geckodriver.exe")
        driver.get(https[0])
        driver_state = 0
    else:
        driver = webdriver.Firefox(executable_path="/mnt/d/geckodriver.exe")
        driver.get(https[0])
        driver_state = 0
    name = get_name(vgst,driver,https[0])
    if driver_state == 0:
                    driver.close()
                    driver_state = -1
    return name

def get_name(vgst, driver,http):
    global drivercount
    web = ['master', 'sahi']
    def iris(vgst, driver):
        vendname = ""
        global driver_state
        c = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.ID,"gstinno"))
        c.send_keys(vgst)
        c.send_keys(Keys.RETURN)
        # button = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.ID,"gstin-search-buton"))
        # button.click()
        try:
            name = WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//div[@class='row brdr']/div/div/input")).get_attribute('value')
        except TimeoutException as ex:
            try:
                error = WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//input[@id=gobackpage]/span")).get_attribute('value')
                if error == "Go back to Search GSTN" :
                    vendname = "Invalid GSTIN"
            except TimeoutException as e:
                if driver_state == 0:
                    driver.close()
                    driver_state = -1               
                vendname = temp_driver(vgst)
        if len(vendname) <= 0:
            vendname = name
        if driver_state == 0:    
            WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//a[text()='SEARCH GSTIN']")).click()
        return vendname
    def sahi(vgst, driver):
        global driver_state
        vendname = ""
        c = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.NAME,"gstin"))
        c.send_keys(vgst)
        c.send_keys(Keys.RETURN)
        # button = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.ID,"gstin-search-buton"))
        # button.click()
        try:
            name = WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//div[@class='row']/div/div/div[2]"))
            while name.text == None:
                name = WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//div[@class='row']/div/div/div[2]"))
        except TimeoutException as ex:
            try:
                error = WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//div[@class='text-danger']")).text
                if error == "err-invalid-gstin-format" :
                    vendname = "Invalid GSTIN"
            except TimeoutException as e:
                if driver_state == 0:
                    driver.close()
                    driver_state = -1
                vendname = temp_driver(vgst)
        if len(vendname) <= 0:
            vendname = name.text
        if driver_state == 0:    
            c = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.NAME,"keyword"))
            c.clear()
        if driver_state == 0:
            driver.refresh()
        return vendname
    def know(vgst, driver):
        global driver_state
        vendname = ""
        c = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.NAME,"gstnum"))
        c.send_keys(vgst)
        c.send_keys(Keys.RETURN)
        # button = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.ID,"gstin-search-buton"))
        # button.click()
        try:
            name = WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//table[@class='striped highlight questionlist']/tbody/tr[1]/td[2]"))
        except TimeoutException as ex:
            try:
                error = WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//div[@class='text-danger']")).text
                if error == "err-invalid-gstin-format" :
                    vendname = "Invalid GSTIN"
            except TimeoutException as e:
                if driver_state == 0:
                    driver.close()
                    driver_state = -1
                vendname = temp_driver(vgst)
        if len(vendname) <= 0:
            vendname = name.text
        return vendname
    def findgst(vgst, driver):
        global driver_state
        vendname = ""
        time.sleep(5)
        c = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.NAME,"gstnum"))
        c.send_keys(vgst)
        driver.implicitly_wait(10)
        c.send_keys(Keys.RETURN)
        # button = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.ID,"gstin-search-buton"))
        # button.click()
        try:
            name = WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//td[text()='Trade Name']/following-sibling::td"))
        except TimeoutException as ex:
            try:
                error = WebDriverWait(driver, 10).until(lambda d: d.find_element_by_xpath("//div[@class='text-danger']")).text
                if error == "err-invalid-gstin-format" :
                    vendname = "Invalid GSTIN"
            except TimeoutException as e:
                if driver_state == 0:
                    driver.close()
                    driver_state = -1
                vendname = temp_driver(vgst)
        if len(vendname) <= 0:
            vendname = name.text
        return vendname

    pattern = re.compile('sahi')
    if pattern.search(http):
        return sahi(vgst,driver)
    pattern = re.compile('know')
    if pattern.search(http):
        return know(vgst, driver)
    pattern = re.compile('find')
    if pattern.search(http):
        return findgst(vgst, driver)
    pattern = re.compile('iris')
    if pattern.search(http):
        return iris(vgst, driver)


def dict_flat(data, row = 0):
    temp = {}
    for key in data.keys():
        if isinstance(data[key], list):
            for x in range(len(data[key])):
                temp.update(dict_flat(data[key][x]))
        elif type(data[key]) == dict:
            temp.update(dict_flat(data[key]))
        else:
            temp.update({key : data[key]})
    return temp

def json2dict(filezz):
    global crsr
    dict1 = {}
    row = 0
    global drivercount
    global driver_state
    for i in range(len(filezz)):
        file = r"/mnt/d/recon/" + filezz[i]
        print(file)
        with open(file) as json_file:
            data = (json.load(json_file))
        if 'b2b' in data:
            for index in range(len(data['b2b'])):
                invc = len(data['b2b'][index]['inv'])
                for x in range(invc):
                    itmsc = len(data['b2b'][index]['inv'][x]['itms'])
                    if data['b2b'][index]['ctin'] not in dict1.keys():
                        dict1[data['b2b'][index]['ctin']] = {}
                    for a in range(itmsc):
                        dict1[data['b2b'][index]['ctin']][row + a] = {}
                    for inv in data['b2b'][index]['inv'][x].keys():
                        if isinstance(data['b2b'][index]['inv'][x][inv], list):
                            continue
                        if itmsc > 1:
                            for y in range(itmsc):
                                dict1[data['b2b'][index]['ctin']][row + y].update(data['b2b'][index]['inv'][x]['itms'][y]['itm_det'])
                                dict1[data['b2b'][index]['ctin']][row + y].update({inv : data['b2b'][index]['inv'][x][inv]})
                                #crsr.execute("select gstin,legalname from gst where gstin = ?", (data['b2b'][index]['ctin'], ))
                                #vname = crsr.fetchall()
                                # if not vname:
                                    # if driver_state == -1:
                                    #     driver = webdriver.Firefox(executable_path="/mnt/d/geckodriver.exe")
                                    #     driver.get(https[0])
                                    #     driver_state = 0
                                    #gstin = (data['b2b'][index]['ctin'] , get_name(data['b2b'][index]['ctin'], driver,https[0]))
                                    #crsr.execute('insert into gst (gstin, legalname) values (?,?)',gstin)
                                    #crsr.execute("select gstin,legalname from gst where gstin = ?", (data['b2b'][index]['ctin'],))
                                    #vname = crsr.fetchall()
                                dict1[data['b2b'][index]['ctin']][row + y].update( {'ctin' : data['b2b'][index]['ctin'] })
                                dict1[data['b2b'][index]['ctin']][row + y].update( {'cname' : 'vlookup' })
                        else:
                            dict1[data['b2b'][index]['ctin']][row].update(dict_flat(data['b2b'][index]['inv'][x]))
                            dict1[data['b2b'][index]['ctin']][row].update({'ctin' : data['b2b'][index]['ctin'] })
                            # crsr.execute("select gstin,legalname from gst where gstin = ?", (data['b2b'][index]['ctin'],))
                            # vname = crsr.fetchall()
                            # if not vname:
                                # print('in if')
                                #if driver_state == -1:
                                    #driver = webdriver.Firefox(executable_path="/mnt/d/geckodriver.exe")
                                    #driver.get(https[0])
                                    #driver_state = 0
                                #gstin = (data['b2b'][index]['ctin'] , get_name(data['b2b'][index]['ctin'], driver,https[0]))
                                #crsr.execute('insert into gst (gstin, legalname) values (?,?)',gstin)
                                #crsr.execute("select gstin,legalname from gst where gstin = ?", (data['b2b'][index]['ctin'],))
                                #vname = crsr.fetchall()
                            dict1[data['b2b'][index]['ctin']][row].update({'cname' : 'vlookup' })
                    row += itmsc
    # print(dict1)
    if driver_state == 0:
        driver.close()
        driver_state = -1
    for keys in dict1.keys():
        for index in dict1[keys].keys():
            if 'iamt' not in dict1[keys][index].keys():
                dict1[keys][index].update({check_vals('iamt') : 0})
            if 'camt' not in dict1[keys][index].keys():
                dict1[keys][index].update({check_vals('camt') : 0})
            if 'samt' not in dict1[keys][index].keys():
                dict1[keys][index].update({check_vals('samt') : 0})
            if 'csamt' not in dict1[keys][index].keys():
                dict1[keys][index].update({check_vals('csamt') : 0})
            for key in dict1[keys][index].keys():
                columns = check_vals(key)
                if columns > -1:
                    dict1[keys][index].update({check_vals(key) : dict1[keys][index][key]})
    for keys in dict1.keys():
        for index in dict1[keys].keys():
            for key in dict1[keys][index].keys():
                columns = check_rem(key)
                if columns == -1:
                    dict1[keys][index].pop(key)
    # print(json.dumps(dict1, indent=4))
    return dict1
# learn delegate
def regexp(inv1, inv2):
    points = 0
    # print(1, inv1, inv2)
    if inv1 == inv2:
        return True
    pattern = re.compile(r"[a-zA-Z]")
    inv1 = pattern.split(inv1)
    inv2 = pattern.split(inv2)
    # print(2, inv1, inv2)
    def ftr(data):
        data1 = []
        for y in range(len(data)):
            if data[y] != '-' and data[y] != '/' and data[y] != '' and data[y] != ' ' and data[y] != '\\':
                data1.append(data[y])
        return data1
    inv1 = list(ftr(inv1))
    inv2 = list(ftr(inv2))
    def jnstr(data):
        data1 = ''
        for x in range(len(data)):
            data1 += data[x]
        return data1
    inv1 = str(jnstr(inv1))
    inv2 = str(jnstr(inv2))
    # print(3, inv1, inv2)
    # def pat(data):
    #     patterns = [r'\d{4}-\d{2}$', r'\d{2}-\d{2}$', r'\/\d{4}-\d{2}\/', r'\/\d{2}-\d{2}\/', r'^\d{4}-\d{2}\/', r'^\d{2}-\d{2}\/']
    #     for x in range(len(patterns)):
    #         pat1 = re.compile(patterns[x])
    #         data = pat1.split(data)
    #         data = str(jnstr(data))
    #     # print('spl', data)
    #     return data
    # inv1 = str(pat(inv1))
    # inv2 = str(pat(inv2))
    # print(4, inv1, inv2)
    if len(inv1) == 0 or len(inv2) == 0:
        return False
    # inv1 = str(jnstr(inv1))
    # inv2 = str(jnstr(inv2))
    # print(5, inv1, inv2)
    # def jn(data):
    #     data1 = 0
    #     for x in range(len(data)):
    #         data1 += int(data[x])
    #     return data1
    # pattern = re.compile(r"\D")
    # inv1 = pattern.split(inv1)
    # inv2 = pattern.split(inv2)
    # inv1 = list(ftr(inv1))
    # inv2 = list(ftr(inv2))
    # print(6, inv1, inv2)
    # inv1 = str(jn(inv1))
    # inv2 = str(jn(inv2))
    # print(7, inv1, inv2)
    if inv1 == inv2:
        return True
    return False

def dfinder(booker):
    ilist = {}
    tdict = dict(booker)
    for gstno in booker.keys():
        ilist[gstno] = {}
        for indexes in sorted(booker[gstno].keys()):
            temp = dict(booker[gstno][indexes])
            for indexes2 in tdict[gstno].keys():
                if indexes == indexes2:
                    continue
                if temp['GSTIN'] == booker[gstno][indexes]['GSTIN'] and regexp(str(temp['Invoice Number']) , str(booker[gstno][indexes2]['Invoice Number'])) and temp['Tax Rate'] == booker[gstno][indexes2]['Tax Rate']:
                    ilist[gstno][indexes] = {"DUPL" : 0}
                    ilist[gstno][indexes2] = {"DUPL" : 0}
    # for gstno in booker.keys():
    #     ilist[gstno] = {}
    #     print(gstno,booker[gstno])
    #     for indexes in sorted(booker[gstno].keys()):
    #         temp = dict(booker[gstno][indexes])
    #         keys = booker[gstno].keys()
    #         keys.sort()
    #         print(keys)
    #         print(indexes)
    #         print(keys[-1])
    #         for key in range(indexes+1, keys[-1] + 1):
    #             if key == keys[-1]:
    #                 continue
    #             if temp['GSTIN'] == booker[gstno][key]['GSTIN'] and regexp(str(temp['Invoice Number']) , str(booker[gstno][key]['Invoice Number'])) and temp['Tax Rate'] == booker[gstno][key]['Tax Rate']:
    #                 ilist[gstno][indexes] = {"DUPL" : 0}
    #                 ilist[gstno][key] = {"DUPL" : 0}
    return ilist

# def valdiff(value1,value2):
#     value1 = round(value1)
#     value2 = round(value2)
#     if value1 == value2:
#         return False
#     elif value1-1 == value2 or value1 == value2-1:
#         return False
#     else:
#         return True

def curate_book(book1, book2,d):
    temp1 = {}
    index = {}
    def valdiff(value1,value2):
        value1 = round(value1)
        value2 = round(value2)
        if value1 == value2:
            return True
        elif value1-1 == value2 or value1 == value2-1:
            return True
        else:
            return False
    def absolute(temp1, temp2):
        count = 0
        for keys in temp1.keys():
            if keys == "CGST Amount" or keys == "SGST Amount" or keys == "IGST Amount" or keys == "Cess Amount":
                if valdiff(temp1[keys], temp2[keys]):
                    count += 1
            elif temp1[keys] == temp2[keys]:
                count += 1
        if count == 14:
            return True
        return False
    mlist1=[]
    mlist2=[]
    for key in book1:
        temp1 = dict(book1[key])
        if key in book2.keys():
            temp2 = dict(book2[key])
        else:
            temp2 = 0
            index[key] = {}
            for ind in book1[key].keys():
                index[key][ind] = { "MISSING BILL" : 0}
        if temp2 != 0:
            for i in temp1.keys():
                for j in temp2:
                    if j in mlist2 or i in mlist1:
                        continue
                    match = absolute(temp1[i], temp2[j])
                    if match:
                        mlist1.append(i)
                        mlist2.append(j)
    def resonable(temp1, temp2):
        count = 0
        for keys in temp1.keys():
            if keys == "CGST Amount" or keys == "SGST Amount" or keys == "IGST Amount" or keys == "Cess Amount" or keys == "Taxable Value" or keys == "Invoice Value":
                if valdiff(temp1[keys], temp2[keys]):
                    count += 1
        if count == 6:
            return True
        return False
    rlist1 = []
    rlist2 = []
    for key in book1:
        temp1 = dict(book1[key])
        if key in book2.keys():
            temp2 = dict(book2[key])
        elif key in index.keys():
            temp2 == 0
            continue
        if temp2 != 0:
            for i in temp1.keys():
                for j in temp2:
                    if j in mlist2 or i in mlist1 or j in rlist2 or i in rlist1:
                        continue
                    match = resonable(temp1[i], temp2[j]) 
                    if match:
                        rlist1.append(i)
                        rlist2.append(j)
    tbooks = 0
    for key in book1:
        for indexx in book1[key].keys():
            if indexx in mlist1 or indexx in rlist1: 
                tbooks += book1[key][indexx]['Taxable Value']
    t2a = 0
    for key in book2:
        for indexx in book2[key].keys():
            if indexx in mlist2 or indexx in rlist2:
                t2a += book2[key][indexx]['Taxable Value']   
    # ilist = dict(dfinder(book1))
    # ilist = { k:v for k, v in ilist.items() if v }
    # for key in ilist.keys():
    #     if key not in index.keys():
    #         index[key] = {}
    #     for ind in ilist[key].keys():
    #         if ind in index[key].keys():
    #             index[key][ind].update(ilist[key][ind])
    #         else:
    #             index[key][ind] = ilist[key][ind]
    for key in book1.keys():
        if key not in index.keys():
            index[key] = {}
        for indie in book1[key].keys():
            if indie in mlist1 or indie in rlist1 or indie in index[key].keys():
                continue
            index[key][indie] = {'MANINT' : 0}
    return { k:v for k, v in index.items() if v }

def ws_header(wsheet):
    for x in range(len(dict_keys)):
        ls = dict_keys[x].keys()
        wsheet.cell(row = 1, column = x + 1 ).value = dict_keys[x][ls[0]]

def mod_excel(wsheet, indexes, leng, name, output):
    redFill = PatternFill(fgColor=RED, fill_type = "solid")
    error = "BILL NOT IN " + name
    indkey = indexes.keys()
    for gstno in output:
        for key in output[gstno].keys():
            for keys in output[gstno][key].keys():
                columns = check_keys(keys)
                if columns > -1:
                    if columns == 3:
                        datee = datetime.strptime(output[gstno][key][keys], '%d-%m-%Y').date()
                        wsheet.cell(row = key + 2, column = columns + 1).value = datee
                    else:
                        wsheet.cell(row = key + 2, column = columns + 1).value = output[gstno][key][keys]
            if gstno in indkey :
                if key in indexes[gstno].keys():
                    if indexes[gstno][key].has_key("MANINT"):
                        wsheet.cell(row = key + 2, column = leng + 1).value = "MANUAL INTERVENTION NEEDED."
                        for indice in range(1,leng+1):
                            wsheet.cell(row = key + 2, column = indice).fill = redFill
                    if indexes[gstno][key].has_key("DUPL"):
                        wsheet.cell(row = key + 2, column = leng + 2).value = "DUPLICATE ENTRY"
                    if indexes[gstno][key].has_key("MISSING BILL"):
                        for indice in range(1,leng+1):
                            if indice == leng:
                                wsheet.cell(row = key + 2, column = indice).fill = redFill
                                wsheet.cell(row = key + 2, column = indice + 3).value = error
                            else:
                                wsheet.cell(row = key + 2, column = indice).fill = redFill


def main():
    global drivercount
    drivercount = 0
    global https
    global broken
    broken = []
    #'https://www.mastersindia.co/gst-number-search-and-gstin-verification/',
    https = ['https://irisgst.com/irisperidot/','https://findgst.in/','https://www.mastersindia.co/gst-number-search-and-gstin-verification/', 'https://app.sahigst.com/search-taxpayer','https://www.knowyourgst.com/gst-number-search/'] 
    jsons = []
    directory = r"/mnt/d/recon/"
    print('connection')
    connection = sqlite3.connect("/mnt/d/recon/gstin.db")
    global crsr
    crsr = connection.cursor()
    crsr.execute('CREATE TABLE IF NOT EXISTS gst(gstin VARCHAR(15) PRIMARY KEY, LEGALNAME VARCHAR(255))')
    files = list_files(directory, "json")
    for f in files:
        jsons.append(f)
    books = []
    twoa = []
    for f in jsons:
        pattern = re.compile(r"offline")
        out = pattern.findall(f)
        if len(out) > 0:
            books.append(f)
        else:
            twoa.append(f)
    global dict_keys
    # print(books)
    dict_keys = { 
                 0 : {'ctin' : 'GSTIN'}, 
                 1 : {'cname' : 'Name'}, 
                 7 : {'val' : 'Invoice Value'},
                 12: {'samt' : 'SGST Amount'}, 
                 11: {'camt' : 'CGST Amount'}, 
                 10: {'iamt' : 'IGST Amount'}, 
                 9 : {'txval' : 'Taxable Value'}, 
                 2 : {'inum' : 'Invoice Number'}, 
                 13: {'csamt' : 'Cess Amount'}, 
                 6 : {'inv_typ' : 'Invoice Type'}, 
                 4 : {'pos' : 'Place of Supply'}, 
                 3 : {'idt' : 'Invoice Date'}, 
                 5 : {'rchrg' : 'Is Reverse Charge Applicaple'}, 
                 8 : {'rt' : 'Tax Rate'},
                 14: {'invrr' : 'Error'},
                 15: {'Dup' : 'Err of Duplication'},
                 16: {'Miss': 'Missing Bill'},
                 }
    global driver_state
    driver_state = -1
    books_dict = json2dict(books)
    print(json.dumps(books_dict))
    twoa_dict = json2dict(twoa)
    # print(json.dumps(twoa_dict))
    connection.commit()
    index_books = dict(curate_book(books_dict,twoa_dict,1))
    index_twoa = dict(curate_book(twoa_dict,books_dict,2))
    books_len = 14
    twoa_len = 14
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "B2B - Ratewise"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "B2B - Ratewise"
    ws_header(ws1)
    ws_header(ws2)
    mod_excel(ws1, index_books, books_len, "GSTR2A", books_dict)
    mod_excel(ws2, index_twoa, twoa_len, "BOOKS", twoa_dict)
    wb1.save(r"/mnt/d/recon/RECONCILED_BOOKS.xlsx")
    wb2.save(r"/mnt/d/recon/RECONCILED_GSTR2A.xlsx")

if __name__ == "__main__":
    main()